# -*- coding: utf-8 -*-
"""Corrige el calendario del Mundial 2026 con el fixture OFICIAL (sorteo del
5-dic-2025): reescribe la hoja PARTIDOS del Excel y resiembra la BD.

Horarios convertidos a hora de Chile (junio = UTC-4).
Fuente: Wikipedia "2026 FIFA World Cup Group A..L" (consultado 12-jun-2026).

Uso: python fix_fixture.py
"""
from datetime import datetime
import sqlite3
import openpyxl

EXCEL_PATH = "DOC-20260611-WA0002..xlsx"
DB_PATH = "prode.db"

# (fecha/hora CL, local, visita, grupo, goles_local, goles_visita)
FIXTURE = [
    ("2026-06-11 15:00", "México", "Sudáfrica", "A", 2, 0),
    ("2026-06-11 22:00", "Corea del Sur", "República Checa", "A", 2, 1),
    ("2026-06-12 15:00", "Canadá", "Bosnia y Herzegovina", "B", None, None),
    ("2026-06-12 21:00", "EEUU", "Paraguay", "D", None, None),
    ("2026-06-13 15:00", "Catar", "Suiza", "B", None, None),
    ("2026-06-13 18:00", "Brasil", "Marruecos", "C", None, None),
    ("2026-06-13 21:00", "Haití", "Escocia", "C", None, None),
    ("2026-06-14 00:00", "Australia", "Turquía", "D", None, None),
    ("2026-06-14 13:00", "Alemania", "Curazao", "E", None, None),
    ("2026-06-14 16:00", "Países Bajos", "Japón", "F", None, None),
    ("2026-06-14 19:00", "Costa de Marfil", "Ecuador", "E", None, None),
    ("2026-06-14 22:00", "Suecia", "Túnez", "F", None, None),
    ("2026-06-15 12:00", "España", "Cabo Verde", "H", None, None),
    ("2026-06-15 15:00", "Bélgica", "Egipto", "G", None, None),
    ("2026-06-15 18:00", "Arabia Saudita", "Uruguay", "H", None, None),
    ("2026-06-15 21:00", "Irán", "Nueva Zelanda", "G", None, None),
    ("2026-06-16 15:00", "Francia", "Senegal", "I", None, None),
    ("2026-06-16 18:00", "Irak", "Noruega", "I", None, None),
    ("2026-06-16 21:00", "Argentina", "Argelia", "J", None, None),
    ("2026-06-17 00:00", "Austria", "Jordania", "J", None, None),
    ("2026-06-17 13:00", "Portugal", "RD Congo", "K", None, None),
    ("2026-06-17 19:00", "Ghana", "Panamá", "L", None, None),
    ("2026-06-17 20:00", "Inglaterra", "Croacia", "L", None, None),
    ("2026-06-17 22:00", "Uzbekistán", "Colombia", "K", None, None),
    ("2026-06-18 12:00", "República Checa", "Sudáfrica", "A", None, None),
    ("2026-06-18 15:00", "Suiza", "Bosnia y Herzegovina", "B", None, None),
    ("2026-06-18 18:00", "Canadá", "Catar", "B", None, None),
    ("2026-06-18 21:00", "México", "Corea del Sur", "A", None, None),
    ("2026-06-19 15:00", "EEUU", "Australia", "D", None, None),
    ("2026-06-19 18:00", "Escocia", "Marruecos", "C", None, None),
    ("2026-06-19 20:30", "Brasil", "Haití", "C", None, None),
    ("2026-06-19 23:00", "Turquía", "Paraguay", "D", None, None),
    ("2026-06-20 13:00", "Países Bajos", "Suecia", "F", None, None),
    ("2026-06-20 16:00", "Alemania", "Costa de Marfil", "E", None, None),
    ("2026-06-20 20:00", "Ecuador", "Curazao", "E", None, None),
    ("2026-06-21 00:00", "Túnez", "Japón", "F", None, None),
    ("2026-06-21 12:00", "España", "Arabia Saudita", "H", None, None),
    ("2026-06-21 15:00", "Bélgica", "Irán", "G", None, None),
    ("2026-06-21 18:00", "Uruguay", "Cabo Verde", "H", None, None),
    ("2026-06-21 21:00", "Nueva Zelanda", "Egipto", "G", None, None),
    ("2026-06-22 13:00", "Argentina", "Austria", "J", None, None),
    ("2026-06-22 17:00", "Francia", "Irak", "I", None, None),
    ("2026-06-22 20:00", "Noruega", "Senegal", "I", None, None),
    ("2026-06-22 23:00", "Jordania", "Argelia", "J", None, None),
    ("2026-06-23 13:00", "Portugal", "Uzbekistán", "K", None, None),
    ("2026-06-23 16:00", "Inglaterra", "Ghana", "L", None, None),
    ("2026-06-23 19:00", "Panamá", "Croacia", "L", None, None),
    ("2026-06-23 22:00", "Colombia", "RD Congo", "K", None, None),
    ("2026-06-24 15:00", "Suiza", "Canadá", "B", None, None),
    ("2026-06-24 15:00", "Bosnia y Herzegovina", "Catar", "B", None, None),
    ("2026-06-24 18:00", "Escocia", "Brasil", "C", None, None),
    ("2026-06-24 18:00", "Marruecos", "Haití", "C", None, None),
    ("2026-06-24 21:00", "República Checa", "México", "A", None, None),
    ("2026-06-24 21:00", "Sudáfrica", "Corea del Sur", "A", None, None),
    ("2026-06-25 16:00", "Curazao", "Costa de Marfil", "E", None, None),
    ("2026-06-25 16:00", "Ecuador", "Alemania", "E", None, None),
    ("2026-06-25 19:00", "Japón", "Suecia", "F", None, None),
    ("2026-06-25 19:00", "Túnez", "Países Bajos", "F", None, None),
    ("2026-06-25 22:00", "Turquía", "EEUU", "D", None, None),
    ("2026-06-25 22:00", "Paraguay", "Australia", "D", None, None),
    ("2026-06-26 15:00", "Noruega", "Francia", "I", None, None),
    ("2026-06-26 15:00", "Senegal", "Irak", "I", None, None),
    ("2026-06-26 20:00", "Cabo Verde", "Arabia Saudita", "H", None, None),
    ("2026-06-26 20:00", "Uruguay", "España", "H", None, None),
    ("2026-06-26 23:00", "Egipto", "Irán", "G", None, None),
    ("2026-06-26 23:00", "Nueva Zelanda", "Bélgica", "G", None, None),
    ("2026-06-27 17:00", "Panamá", "Inglaterra", "L", None, None),
    ("2026-06-27 17:00", "Croacia", "Ghana", "L", None, None),
    ("2026-06-27 19:30", "Colombia", "Portugal", "K", None, None),
    ("2026-06-27 19:30", "RD Congo", "Uzbekistán", "K", None, None),
    ("2026-06-27 22:00", "Argelia", "Austria", "J", None, None),
    ("2026-06-27 22:00", "Jordania", "Argentina", "J", None, None),
]


def build_rows():
    rows = []
    matches = sorted(FIXTURE, key=lambda m: m[0])
    assert len(matches) == 72, len(matches)
    now = datetime.now()
    for i, (fecha, local, visita, grupo, gh, ga) in enumerate(matches, start=1):
        dt = datetime.strptime(fecha, "%Y-%m-%d %H:%M")
        estado = "Cerrado" if dt <= now else "Abierto"
        resultado = f"{gh}-{ga}" if gh is not None else None
        rows.append((f"M{i:02d}", dt, local, visita, f"Grupo {grupo}",
                     gh, ga, resultado, estado))
    return rows


def fix_excel(rows):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["PARTIDOS"]
    # Limpiar filas de datos existentes (deja la fila 1 de encabezados).
    ws.delete_rows(2, ws.max_row)
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            if c == 2:
                cell.number_format = "DD/MM/YYYY HH:MM"
    wb.save(EXCEL_PATH)
    print(f"[OK] Excel actualizado: {len(rows)} partidos en hoja PARTIDOS.")


def fix_db(rows):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    # El fixture anterior era inventado: se borran partidos y pronósticos viejos.
    n_pred = cur.execute("SELECT COUNT(*) FROM predictions").fetchone()[0]
    cur.execute("DELETE FROM predictions")
    cur.execute("DELETE FROM matches")
    conn.commit()
    conn.close()
    print(f"[OK] BD limpiada ({n_pred} pronósticos del fixture viejo eliminados).")

    from seed_db import populate_matches
    populate_matches()

    # Cargar resultados ya jugados y recalcular ranking.
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    for mid, _dt, _l, _v, _g, gh, ga, _res, _e in rows:
        if gh is not None:
            cur.execute(
                "UPDATE matches SET home_goals=?, away_goals=?, status='FT' WHERE id=?",
                (gh, ga, mid))
    conn.commit()
    conn.close()
    from main import recalculate_scores
    recalculate_scores()
    print("[OK] Resultados jugados cargados y ranking recalculado.")


if __name__ == "__main__":
    rows = build_rows()
    fix_excel(rows)
    fix_db(rows)
